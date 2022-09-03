from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
import openpyxl
from getpass import getpass
import time
import re



class LinkedIn():
    wb=Workbook()
    ws=wb.active
    ws['A1']='Job Title'
    ws['B1']='Organisation'
    ws['C1']='Location'
    ws['D1']='Job URL'
    email=input('Enter your Login Email: ')
    password=getpass("Enter your password: ")
    search_term=input("Enter search term: ")
    
    
    
    
    ausnz_keyword=['australian citizen','Australian citizen','Australian Citizen','australian citizenship','Australian Citizenship','Australian citizenship' 
,'australian/nz Citizen','australian/nz citizen','Australian/NZ Citizen','Australian/NZ citizen',
'Australian/NZ citizenship','Australian/NZ Citizenship', 'australian/nz Citizenship', 'australian/nz citizenship',
'New Zealand citizen','New Zealand Citizen','new zealand citizen','new zealand Citizen',
'New Zealand Citizenship','new zealand Citizenship','New Zealand citizenship', 'new zealand citizenship',
'Australian PR','Australian pr','australian PR','permanent resident','permanent residency','Permanent Resident','Permanent Residency', 'nv1', 'security clearance', 'military', 'defence force', 'agsva']


    visa_keywords=['485', 'valid work visa', 'temporary resident', 'working rights', 'work rights','Valid Work Visa', 
    'work visa','Valid work visa', 'Temporary resident', 'Working rights', 'Work Rights', 'Work visa','Visa','visa',
    'Work Visa','VALID WORK VISA', 'TEMPORARY RESIDENT', 'WORKING RIGHTS', 'WORK RIGHTS', 'WORK VISA', 
    'VALID WORK VISA', 'TEMPORARY RESIDENT', 'WORKING RIGHTS', 'WORK RIGHTS', 'WORK VISA']
    Job_URL=[]
    job_title=[]
    job_url_id=[]
    duplicate_job_id=[]
    final_unique_id_list=[]
    job_loc=[]
    organ_name=[]
    scroll_counter=0
    
    def __init__(self):
        self.driver=webdriver.Chrome(executable_path=ChromeDriverManager().install())   #install driver from path
        

    def Login(self):
        driver=self.driver
        driver.get("https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin") #Open the URL    
        
        #Login Page input credentials
        
        login_email=driver.find_element_by_id('username')
        login_email.send_keys(self.email)
        login_pwd=driver.find_element_by_id('password')
        login_pwd.send_keys(self.password)
        login_pwd.send_keys(Keys.RETURN)
    
    def login_check(self):
        driver=self.driver
        driver.maximize_window()
        self.Login()
        time.sleep(15)
        if driver.find_element_by_xpath("//li-icon[@aria-hidden='true'][@type='nav-small-home-icon']").is_displayed()==True:
            
            time.sleep(3)
            self.move_to_jobs()
        else:
            time.sleep(15)
            print('\n You have to verify')
            self.move_to_jobs()

    def move_to_jobs(self):
        
        driver=self.driver
        time.sleep(5)
        
        print('\n Move to jobs')
        # Jobs Page
        driver.get('https://www.linkedin.com/jobs/search/')
        time.sleep(1)
        
        #Putting the skill in the searchbox
        c=driver.find_element_by_xpath("//input[@aria-label='Search by title, skill, or company']")
        c.send_keys(search_term)
        time.sleep(1)
        d=driver.find_element_by_xpath("//button[normalize-space()='Search']").click()   #Pressing Search Button
        
        time.sleep(3)
        
        driver.find_element_by_xpath("//li-icon[@type='chevron-down'][@class='artdeco-button__icon']").click()
        
        time.sleep(2)


        
        if input('''\n \n \n Do you want to add experience filter?: (Y/N)
PLEASE NOTE: If you want to see openings for all experience level press 'N' \n''').upper()=='Y':
            print('\n add_experience_filter inititated first')
            self.add_experience_filter()

        else:
            print('\n Innitiating next page')
            self.next_page()
        
        time.sleep(2)
        

    def add_experience_filter(self):
        # self.move_to_jobs()
        driver=self.driver

        a=driver.find_element_by_xpath("//label[contains(@for,'experience-1')]//span[contains(@class,'t-14 t-black--light t-normal')][normalize-space()='Internship']")
        b=driver.find_element_by_xpath("//label[@for='experience-2']")
        c=driver.find_element_by_xpath("//label[contains(@for,'experience-3')]")
        d=driver.find_element_by_xpath("//label[@for='experience-4']")
        e=driver.find_element_by_xpath("//label[contains(@for,'experience-5')]")
        f=driver.find_element_by_xpath("//label[contains(@for,'experience-6')]")



        driver.find_element_by_xpath("//button[normalize-space()='Experience Level']").click()
        print('''Which experience level best suits you:
a. Internship
b. Entry Level
c. Associate
d. Mid-Senior Level
e. Director
f. Executive''')
        user_exp=input("Enter the corresponding letter next to your experience level (add comma if more than 1)")
        user_exp_list=user_exp.split(',')
        print(user_exp_list)
        #Add Filter by Experience Level
        for i in user_exp_list:
            if i not in ['a','b','c','d','e','f']:
                print('Invalid letter')
                break
            elif i=='a':
                a.click()
            elif i=='b':
                b.click()
            elif i=='c':
                c.click()
            elif i=='d':
                d.click()
            elif i=='e':
                e.click()
            elif  i=='f':
                f.click()
        


                
        driver.find_element_by_xpath("//header[@id='global-nav']").click()
        time.sleep(5)
        print('\n Initiating Next Page Module')

        self.next_page()
            

    def next_page(self):
        driver=self.driver
        
        res_container=25
        tot_page=driver.find_element_by_xpath("//ul[@class='artdeco-pagination__pages artdeco-pagination__pages--number']//li[10]").text
        total_page=int(tot_page)
        tot_res=driver.find_element_by_xpath("//div[@class='jobs-search-results-list__title-heading']/small").text
        print('\n\n There are {0} based on your search term and experience preferences over {1} pages'.format(tot_res, tot_page))
        page_counter=int(input('\n How many pages of results do you want to go through? '))
        linx=driver.current_url
        
        # if page_counter<=total_page:
        for i in range(1,page_counter):
            
            print('Initiating Scroll and load', self.scrolling_and_getting_ID())
            time.sleep(5)
            print('Initiating Scrape and Append',self.scrape_keywords())
            URL=driver.current_url
            URL_manipulate=URL+'&start={}'.format(res_container)
            print('\n JUMPING TO NEXT PAGE!!')
            driver.get(URL_manipulate)
            res_container+=25
            print('\n append list:', self.job_title)
            print('\n append list length:', len(self.job_title))
            time.sleep(10)
        self.excel_entry()
        time.sleep(60)
        
    def scrolling_and_getting_ID(self):
        ## Scrolling
        
        ##Using while loop to scroll a couple of times to load all the web elements in the page
        ##The while loop also takes in all the job ID from the left pane of the page and append in Job_URL
                       
        driver=self.driver
        while self.scroll_counter<15:
            
            list1=driver.find_elements_by_class_name('job-card-container')              # Selecting the job container
            # print(len(list1))                                                         ## Enable this to see the initial length of the list of job container

            job_id=list1[-1].get_attribute('data-job-id')                               # Storing the last JobID generated at initial loadup in a variable
            xpath_1="//div[@data-job-id='{}']".format(job_id)                           # Again formatting a XPath with the unique JobID
            # print("\n The xpath1 link is: ", xpath_1)                                 ## Enable to see the JobID XPath
            
            
            scrolling_click_link=driver.find_element_by_xpath(xpath_1)                  # Detecting the created XPath and storing in a variable to create a reference for scrolling
            for items in list1:                                                         # The for loop moves through "List1" (The job container)
                self.Job_URL.append(items.get_attribute('data-job-id'))                 # Storincg all the JobID generated in a list "Job_URL"

            driver.execute_script("arguments[0].scrollIntoView(true);", scrolling_click_link)   #Enable Scrolling through Javascript executer while refering to the JobID XPath stored in "scrolling_click_link"

            job_id=list1[-1].get_attribute('data-job-id')                               # Replacing "job_id" value with the new value of the last element in the "job container list"
            # print("\n the job id after iteration is: ", job_id)                       ## Enable to see the new 'job_id' value

            self.scroll_counter+=1                                                      # Increment the scroll counter
            
        self.scroll_counter=0
        


        #Removing Duplicates from Job_URL list
        for i in self.Job_URL:
            if i not in self.final_unique_id_list:
                self.final_unique_id_list.append(i)
            else:
                self.duplicate_job_id.append(i)
        
 

        self.Job_URL.clear()                                                            #Reusing the same Job_URL
        
        #Proofing the list to get all the job_id on the page (25 results/page)
        list2=driver.find_elements_by_class_name('job-card-container')
        for i in list2:
            self.Job_URL.append(i.get_attribute('data-job-id'))

        print('\n')
        print('\n')
        
        


    def scrape_keywords(self):
    #Getting links
        driver=self.driver
        
        citizen_kw_match_counter=0
        citizen_kw=[]
        visa_kw=[]
        
        no_citi_no_visa=0
        list_keyword_counter=0
        vkeyword_counter=0
        for i in self.Job_URL:
            xpath_job_id="//div[@data-job-id='{}']".format(i)
            get_link=driver.find_element_by_xpath(xpath_job_id).click()
            job_title=driver.find_element_by_xpath("//h2[@class='t-24 t-bold jobs-unified-top-card__job-title']").text
            job_description=driver.find_element_by_xpath("//div[@id='job-details']").text
            org_name=driver.find_element_by_xpath("//span[@class='jobs-unified-top-card__company-name']").text
            job_location=driver.find_element_by_xpath("//span[@class='jobs-unified-top-card__bullet']").text
            
            job_description=job_description.lower()

            if re.search(r'citizen', string=job_title.lower()):
                for title in self.ausnz_keyword:
                    if re.search(title, string=job_title.lower()):
                        citizen_kw_match_counter+=1
                        
            else:
                for visa in self.visa_keywords:

                    if re.search(visa, string=job_description):
                        linx=driver.current_url
                        self.job_title.append(job_title)
                        self.job_url_id.append(linx)
                        self.job_loc.append(job_location)
                        self.organ_name.append(org_name)
                        vkeyword_counter+=1
                        
            
                    elif re.search(r'citizen', string=job_description):
                        kw=re.compile(r'citizen')
                        kw_match=kw.finditer(job_description)
                        for match in kw_match:
                            for j in self.ausnz_keyword:
                                if re.search(j, string=job_description):
                                    citizen_kw_match_counter+=1
                                    break
                            break
                    
                    else:
                        linx=driver.current_url
                        self.job_title.append(job_title)
                        self.job_url_id.append(linx)
                        self.job_loc.append(job_location)
                        self.organ_name.append(org_name)
                        no_citi_no_visa+=1

                    break
        time.sleep(1)


       
   
        
            
    def excel_entry(self):
        for index, value in enumerate(self.job_title):
            self.ws.cell(row=index+2, column=1, value=value)
        for index, value in enumerate(self.organ_name):
            self.ws.cell(row=index+2, column=2, value=value)
        for index, value in enumerate(self.job_loc):
            self.ws.cell(row=index+2, column=3, value=value)
        for index, value in enumerate(self.job_url_id):
            self.ws.cell(row=index+2, column=4, value=value)
        self.wb.save('D:\Python coding\Selenium\Selenium Projects\Job_Links.xlsx')
        print('\n \n Jobs pertaining to your keywords and experience have been added to the excel workbook.')

            

        
        
        

    

        
        
        

findr=LinkedIn()
findr.login_check()
        
  
            
    
      
 
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
